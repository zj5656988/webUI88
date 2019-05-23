from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl import worksheet
from commonfunction.phone_no import PhoneNo

class Excel():
    def excel_read(self,str):
        wb = load_workbook("C:\\Users\Administrator\Desktop\\web_UI.xlsx")
        case_shell = wb["case"]
        data=case_shell[str].value
        return data


    def excel_write(self,list):
        wb=load_workbook("C:\\Users\Administrator\Desktop\\web_UI.xlsx")
        data_sheet=wb["data"]
        # data_sheet.cell[]


if __name__ == '__main__':
    m=PhoneNo().getno()
    excel=Excel()
    print(m[0])
    excel.excel_write(m)








